import sys
from pathlib import Path
import openpyxl
from app import create_app
from app.db import create_martyr, get_colleges, get_weapons_by_college, init_db, seed_admin, seed_reference_data


def norm_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def norm_date(value):
    if value is None:
        return None
    if hasattr(value, 'strftime'):
        return value.strftime('%Y-%m-%d')
    text = str(value).strip()
    return text or None


def find_college_id(app, raw_name):
    if not raw_name:
        return None
    normalized = str(raw_name).replace('_', ' ').strip()
    for row in get_colleges(app):
        if row['name_ar'] in normalized or normalized in row['name_ar']:
            return row['id']
    aliases = {
        'الكلية الحربية': 1,
        'الكلية الفنية العسكرية': 2,
        'الكلية الجوية': 3,
        'الكلية البحرية': 4,
        'الدفاع الجوى': 5,
        'الدفاع الجوي': 5,
        'الكلية التكنولوجية العسكرية': 6,
    }
    return aliases.get(normalized)


def find_weapon_id(app, college_id, raw_weapon):
    if not college_id or not raw_weapon:
        return None
    target = str(raw_weapon).strip().replace('ى', 'ي')
    for item in get_weapons_by_college(app, college_id):
        name = item['name_ar'].replace('ى', 'ي')
        if target == name or target in name or name in target:
            return item['id']
    return None


def main():
    if len(sys.argv) < 2:
        print('Usage: python scripts/import_martyrs.py <xlsx_path>')
        raise SystemExit(1)
    path = Path(sys.argv[1])
    if not path.exists():
        print(f'File not found: {path}')
        raise SystemExit(1)

    app = create_app()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = 0
    created = 0
    with app.app_context():
        init_db(app)
        seed_reference_data(app)
        seed_admin(app)
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows += 1
            row = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
            full_name = norm_text(row.get('الأسم') or row.get('الاسم'))
            if not full_name:
                continue
            college_id = find_college_id(app, row.get('الكلية'))
            weapon_id = find_weapon_id(app, college_id, row.get('السلاح'))
            sons = int(row.get('عدد الأبناء ( الذكور )') or 0)
            daughters = int(row.get('عدد الأبناء ( أناث )') or 0)
            data = {
                'full_name': full_name,
                'military_rank': None,
                'college_id': college_id,
                'weapon_id': weapon_id,
                'custom_weapon': norm_text(row.get('السلاح')) if weapon_id is None else None,
                'governorate': norm_text(row.get('المحافظة')),
                'birth_date': norm_date(row.get('تاريخ الميلاد')),
                'martyrdom_date': norm_date(row.get('تاريخ الأستشهاد')),
                'age_at_martyrdom': float(row.get('السن عند الأستشهاد') or 0) if row.get('السن عند الأستشهاد') is not None else None,
                'marital_status': norm_text(row.get('الحاله الأجتماعية')),
                'brothers_count': int(row.get('عدد الأخواه') or 0),
                'sisters_count': int(row.get('عدد الأخوات') or 0),
                'sons_count': sons,
                'daughters_count': daughters,
                'children_count': sons + daughters,
                'father_phone': norm_text(row.get('رقم موبيل الوالد')),
                'mother_phone': norm_text(row.get('رقم موبيل الوالده')),
                'alternate_phone': norm_text(row.get('رقم موبيل اخر')),
                'alternate_phone_owner': norm_text(row.get('اسم مالك الرقم الاخر')),
                'family_guardian_name': None,
                'family_phone': norm_text(row.get('رقم موبيل الوالد')) or norm_text(row.get('رقم موبيل الوالده')),
                'family_address': norm_text(row.get('المحافظة')),
                'monthly_support_needed': 0,
                'urgent_need': 0,
                'support_priority': 'normal',
                'family_status': None,
                'notes': 'Imported from martyrs Excel sheet',
                'image_path': None,
                'is_active': 1,
            }
            create_martyr(app, data)
            created += 1
    print(f'Imported {created} martyrs from {rows} rows')


if __name__ == '__main__':
    main()
